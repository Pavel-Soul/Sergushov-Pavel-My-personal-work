import os
import asyncio
from pymongo import MongoClient, ReturnDocument
from datetime import datetime
import json
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

from vk_teams_async_bot.dispatcher import Dispatcher
from vk_teams_async_bot.bot import Bot
from vk_teams_async_bot.events import Event # Event понадобится для аннотации типов в колбэках
from vk_teams_async_bot.handler import MessageHandler, CommandHandler, BotButtonCommandHandler # Добавили BotButtonCommandHandler
from vk_teams_async_bot.helpers import InlineKeyboardMarkup, KeyboardButton
from vk_teams_async_bot.constants import StyleKeyboard, ParseMode
# Filter пока не используется, можно закомментировать или удалить
# from vk_teams_async_bot.filter import Filter

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# --- НАСТРОЙКИ ---
try:
    CHATID = os.environ["TEST_CHATID"]
    API_URL_BASE = os.environ["API_URL_BASE"]
    TOKEN = os.environ["TEST_TOKEN"]
except KeyError as e:
    logger.critical(f"Отсутствует переменная окружения: {e}. Бот не сможет запуститься.")
    exit()

# Подключение к MongoDB
try:
    client = MongoClient('localhost', 27017)
    client.admin.command('ping') # Проверка доступности сервера
    db_roBOD = client['compl']
    logger.info("Успешное подключение к MongoDB.")
except Exception as e:
    logger.critical(f"Не удалось подключиться к MongoDB: {e}. Бот не сможет запуститься.")
    exit()


last_press = {} # Для отслеживания состояния "ожидание ответа"

app = Bot(bot_token=TOKEN, url=API_URL_BASE)
dispatcher = Dispatcher(bot=app) # Диспетчер теперь создается здесь

# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---

def create_keyboard_from_list(buttons_list_of_dicts: list) -> InlineKeyboardMarkup:
    """
    Создает InlineKeyboardMarkup из списка списков словарей, описывающих кнопки.
    Каждый внутренний список словарей представляет ряд кнопок.
    Каждая кнопка в ряду - это словарь с ключами text, callbackData, style.
    """
    keyboard_rows = []
    for row_of_buttons_data in buttons_list_of_dicts:
        buttons_in_row = []
        for button_data in row_of_buttons_data:
            buttons_in_row.append(
                KeyboardButton(
                    text=button_data["text"],
                    callback_data=button_data["callbackData"], # callbackData уже должен быть строкой
                    style=button_data.get("style", StyleKeyboard.BASE)
                )
            )
        keyboard_rows.append(buttons_in_row)
    return InlineKeyboardMarkup(inline_keyboard=keyboard_rows)


def get_ReqText_from_db(ReqNumb: int):
    """Получает текст заявки и информацию о создателе из MongoDB."""
    req = db_roBOD['requests'].find_one({'ReqNumb': ReqNumb})
    if req:
        msgtxt = (f"Cообщение от @[{req['CreatorUserId']}]:\n\n"
                  f"<b>\"{req['ReqText']}\"</b>\n"
                  f"Создана заявка с номером <b>{str(req['ReqNumb'])}</b>.")
        return msgtxt, req # Возвращаем и текст, и саму запись
    else:
        logger.warning(f"Заявка с номером {ReqNumb} не найдена в MongoDB.")
        return None, None

def create_xlsx_report():
    """Создание xls отчета на основе данных из MongoDB."""
    wb = Workbook()
    ws = wb.active
    headers = ['Номер заявки', 'Текст заявки', 'Отправитель', 'Дата отправки', 'Исполнитель', 'Дата начала исполнения', 'Дата выполнения', 'Статус']
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
    try:
        reqs_cursor = db_roBOD['requests'].find({}).sort("ReqNumb", 1)
        for req in reqs_cursor:
            ws.append([
                req.get('ReqNumb'), req.get('ReqText'), req.get('CreatorUserId'),
                req.get('ReqStartDt').strftime('%Y-%m-%d %H:%M:%S') if req.get('ReqStartDt') else '',
                req.get('ReqOperatorUserId'),
                req.get('ReqExecStartDt').strftime('%Y-%m-%d %H:%M:%S') if req.get('ReqExecStartDt') else '',
                req.get('ReqExecEndDt').strftime('%Y-%m-%d %H:%M:%S') if req.get('ReqExecEndDt') else '',
                req.get('Status')
            ])

        # Автоматическая ширина столбцов
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > 0 else 10
            ws.column_dimensions[column_letter].width = adjusted_width
            if headers[col_num-1] == 'Текст заявки':
                 for cell_in_col in ws[column_letter]:
                    cell_in_col.alignment = cell_in_col.alignment.copy(wrap_text=True)

    except Exception as e:
        logger.exception(f"Ошибка при получении данных для отчета из MongoDB: {e}")
        return None

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    report_timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    file_stream.name = f'{report_timestamp}_report.xlsx'
    return file_stream

# --- ОБРАБОТЧИКИ СОБЫТИЙ ---

async def start_command_cb(event: Event, bot: Bot):
    user_id = event.user.id
    last_press.pop(user_id, None) # Сброс состояния
    await bot.send_text(
        chat_id=event.chat.chat_id,
        text="Здравствуйте. Проконсультирую об инсайдерской информации.\nСформулируйте, пожалуйста, вопрос в одном сообщении.🤖"
    )

async def report_command_cb(event: Event, bot: Bot):
    if event.chat.chat_id == CHATID: # Только для основного чата
        logger.info(f"Запрос отчета от {event.user.id} в группе {CHATID}")
        file_stream = create_xlsx_report()
        if file_stream:
            await bot.send_file(
                chat_id=CHATID,
                file=file_stream, # Используем file вместо bytes_io_object или file_stream
                caption="Отчет по заявкам"
            )
            logger.info("Отчет успешно отправлен.")
        else:
            await bot.send_text(chat_id=CHATID, text="Не удалось сформировать отчет.")
    else:
        logger.warning(f"Команда /report из неосновной группы {event.chat.chat_id} от {event.user.id}")


async def message_cb(event: Event, bot: Bot):
    logger.info(f"Сообщение: '{event.text}' от {event.user.id} в чате {event.chat.type} ({event.chat.chat_id})")
    user_id = event.user.id
    user_first_name = event.user.first_name
    chat_id = event.chat.chat_id

    # Пользователь присылает ответ на заявку (после нажатия "Написать/Редактировать ответ")
    user_state = last_press.get(user_id)
    if event.chat.type == "private" and \
       isinstance(user_state, (list, tuple)) and len(user_state) > 0 and \
       user_state[0] == 'Написать/Редактировать ответ':
        try:
            ReqNumb = int(user_state[1])
            group_msg_id_for_closure = user_state[2] # ID сообщения в группе

            db_roBOD['requests'].update_one({'ReqNumb': ReqNumb}, {"$set": {'Answer': event.text}})
            logger.info(f"Ответ для заявки {ReqNumb} сохранен: {event.text}")

            txt = 'Для отправки ответа пользователю нажмите кнопку "Закрыть заявку":'
            markup_list = [[
                {"text": 'Закрыть заявку', "callbackData": f'Закрыть заявку;{ReqNumb};{group_msg_id_for_closure}', "style": StyleKeyboard.PRIMARY}
            ]]
            await bot.send_text(
                chat_id=user_id, # Отправляем оператору
                text=txt,
                parse_mode=ParseMode.HTML,
                inline_keyboard_markup=create_keyboard_from_list(markup_list)
            )
            last_press.pop(user_id, None) # Сбрасываем состояние ожидания ответа
        except Exception as e:
            logger.exception(f"Ошибка при обработке ответа на заявку (ReqNumb из {user_state}): {e}")
            await bot.send_text(chat_id=user_id, text="Произошла ошибка при сохранении вашего ответа.")

    # Новая заявка от пользователя в приватном чате (не команда и не ответ)
    elif event.chat.type == "private" and not (event.text and event.text.startswith('/')):
        try:
            # Атомарное увеличение и получение номера заявки
            update_result = db_roBOD['lastReqNumb'].find_one_and_update(
                {},
                {"$inc": {'ReqNumb': 1}},
                projection={'ReqNumb': 1, '_id': 0},
                upsert=True,
                return_document=ReturnDocument.AFTER
            )
            if not update_result: # Если upsert не сработал (маловероятно)
                db_roBOD['lastReqNumb'].insert_one({'ReqNumb': 1})
                curr_ReqNumb = 1
            else:
                curr_ReqNumb = update_result['ReqNumb']

            # Отправляем сообщение создателю заявки
            txt_to_creator = f"{user_first_name},\nПо Вашему обращению создана заявка с № <b>{curr_ReqNumb}</b>\n"
            sent_message_creator = await bot.send_text(chat_id=chat_id, text=txt_to_creator, parse_mode=ParseMode.HTML)
            creator_msg_id = sent_message_creator.message_id if sent_message_creator else None

            # Определяем текст заявки (текст или файл)
            ReqText = event.text
            if not ReqText and event.message and event.message.parts:
                file_part = event.message.parts[0].payload
                file_id = file_part.file_id
                caption = file_part.caption if file_part.caption else ""
                # Ссылка на файл может отличаться в VK Teams, уточните формат
                ReqText = f'Вложение: https://files.teams.vk.com/get/{file_id}\n{caption}'.strip()
            elif not ReqText:
                 ReqText = 'Сообщение без текста и вложений'

            # Помещаем заявку в БД
            db_roBOD['requests'].insert_one({
                'ReqNumb': curr_ReqNumb,
                'ReqText': ReqText,
                'ReqStartDt': datetime.now(),
                'CreatorUserId': user_id,
                'CreatorFirstName': user_first_name,
                'ReqOperatorUserId': '', # Пусто, т.к. еще не принята
                'ReqExecStartDt': None, # Используем None для дат, которые еще не установлены
                'ReqExecEndDt': None,
                'Status': 'Создана',
                'CreatorMsgId': creator_msg_id,
                'Answer': ''
            })
            logger.info(f"Создана заявка {curr_ReqNumb} от {user_id}: {ReqText[:50]}...")

            # Отправляем сообщение в группу
            msgtxt_for_group, _ = get_ReqText_from_db(curr_ReqNumb)
            if msgtxt_for_group:
                markup_list_group = [[
                    {"text": 'Принять заявку', "callbackData": f'Принять заявку;{curr_ReqNumb}', "style": StyleKeyboard.PRIMARY}
                ]]
                await bot.send_text(
                    chat_id=CHATID,
                    text=msgtxt_for_group,
                    parse_mode=ParseMode.HTML,
                    inline_keyboard_markup=create_keyboard_from_list(markup_list_group)
                )
        except Exception as e:
            logger.exception(f"Ошибка при создании новой заявки от {user_id}: {e}")
            await bot.send_text(chat_id=chat_id, text="Произошла ошибка при создании вашей заявки. Пожалуйста, попробуйте позже.")


async def buttons_answer_cb(event: Event, bot: Bot):
    # Для BotButtonCommandHandler, event - это уже объект CallbackQuery, содержащий нужные данные
    callback_data_str = event.callback_query.data
    operator_user_id = event.callback_query.from_user.id
    # operator_chat_id = operator_user_id # ЛС оператора - это его ID
    message_in_group_id = None # Будет определен ниже, если нужен

    logger.info(f"Нажатие кнопки: '{callback_data_str}' от {operator_user_id}")
    callback_parts = callback_data_str.split(';')
    action = callback_parts[0]

    try:
        if action == 'Принять заявку':
            ReqNumb = int(callback_parts[1])
            # Сообщение, на котором нажали кнопку "Принять", находится в группе
            message_in_group_id = event.callback_query.message.message_id

            req_before_update = db_roBOD['requests'].find_one({'ReqNumb': ReqNumb})

            if not req_before_update:
                logger.warning(f"Попытка принять несуществующую заявку {ReqNumb} оператором {operator_user_id}")
                await bot.answer_callback_query(
                    query_id=event.callback_query.query_id,
                    text=f"Заявка №{ReqNumb} не найдена.",
                    show_alert=True
                )
                return

            if req_before_update.get('ReqOperatorUserId'):
                logger.info(f"Заявка {ReqNumb} уже принята ({req_before_update['ReqOperatorUserId']}), попытка принять от {operator_user_id}")
                await bot.answer_callback_query(
                    query_id=event.callback_query.query_id,
                    text=f"Заявка уже принята @[{req_before_update['ReqOperatorUserId']}].",
                    show_alert=True
                )
                return

            # Обновляем заявку в БД
            db_roBOD['requests'].update_one(
                {'ReqNumb': ReqNumb},
                {"$set": {
                    'ReqOperatorUserId': operator_user_id,
                    'ReqExecStartDt': datetime.now(),
                    'Status': 'В работе'
                }}
            )
            logger.info(f"Заявка {ReqNumb} принята оператором {operator_user_id}")

            # Обновляем сообщение в группе (удаляем кнопки, добавляем инфо об исполнителе)
            msgtxt_orig, _ = get_ReqText_from_db(ReqNumb)
            opertxt = f'\n\nИсполнитель:\n@[{operator_user_id}].'
            await bot.edit_text(
                chat_id=CHATID, # Редактируем в группе
                message_id=message_in_group_id,
                text=msgtxt_orig + opertxt,
                parse_mode=ParseMode.HTML,
                inline_keyboard_markup= InlineKeyboardMarkup(inline_keyboard=[]) # Пустая клавиатура для удаления кнопок
            )

            # Отправляем сообщение оператору с кнопками управления
            markup_operator_controls_list = [
                [{"text": 'Отказаться от заявки', "callbackData": f'Отказаться от заявки;{ReqNumb};{message_in_group_id}', "style": StyleKeyboard.BASE}],
                [{"text": 'Написать/Редактировать ответ', "callbackData": f'Написать/Редактировать ответ;{ReqNumb};{message_in_group_id}', "style": StyleKeyboard.PRIMARY}],
                [{"text": 'Закрыть заявку', "callbackData": f'Закрыть заявку;{ReqNumb};{message_in_group_id}', "style": StyleKeyboard.PRIMARY}]
            ]
            await bot.send_text(
                chat_id=operator_user_id, # Отправляем оператору в ЛС
                text=msgtxt_orig + "\n\nВы приняли эту заявку.",
                parse_mode=ParseMode.HTML,
                inline_keyboard_markup=create_keyboard_from_list(markup_operator_controls_list)
            )

            # Уведомляем создателя заявки
            creator_user_id = req_before_update['CreatorUserId']
            creator_msg_id = req_before_update.get('CreatorMsgId')
            creator_first_name = req_before_update.get('CreatorFirstName', 'Пользователь')

            if creator_user_id and creator_msg_id:
                txt_to_creator_update = (f"{creator_first_name},\n"
                                         f"По Вашему обращению (заявка № <b>{ReqNumb}</b>) назначен исполнитель: @[{operator_user_id}].")
                try:
                    await bot.edit_text(
                        chat_id=creator_user_id,
                        message_id=creator_msg_id,
                        text=txt_to_creator_update,
                        parse_mode=ParseMode.HTML
                    )
                except Exception as e_edit:
                    logger.error(f"Не удалось отредактировать сообщение создателю {creator_user_id} для заявки {ReqNumb}: {e_edit}")
                    # Если редактирование не удалось, можно отправить новое сообщение
                    await bot.send_text(chat_id=creator_user_id, text=txt_to_creator_update, parse_mode=ParseMode.HTML)
            await bot.answer_callback_query(query_id=event.callback_query.query_id, text="Заявка принята.")


        elif action in ('Отказаться от заявки', 'Отпустить заявку'):
            ReqNumb = int(callback_parts[1])
            message_in_group_id = callback_parts[2] # ID сообщения в группе, которое нужно обновить

            db_roBOD['requests'].update_one(
                {'ReqNumb': ReqNumb, 'ReqOperatorUserId': operator_user_id}, # Убедимся, что отказывается тот, кто принял
                {"$set": {
                    'ReqOperatorUserId': '',
                    'ReqExecStartDt': None,
                    'Status': 'Создана',
                    'Answer': '' # Сбрасываем ответ, если был написан, но не отправлен
                }}
            )
            logger.info(f"Оператор {operator_user_id} отказался от заявки {ReqNumb}")

            # Восстанавливаем сообщение в группе с кнопкой "Принять заявку"
            msgtxt_group_reopen, _ = get_ReqText_from_db(ReqNumb)
            if msgtxt_group_reopen:
                markup_group_reopen_list = [[
                    {"text": 'Принять заявку', "callbackData": f'Принять заявку;{ReqNumb}', "style": StyleKeyboard.PRIMARY}
                ]]
                await bot.edit_text(
                    chat_id=CHATID,
                    message_id=message_in_group_id, # ID исходного сообщения в группе
                    text=msgtxt_group_reopen,
                    parse_mode=ParseMode.HTML,
                    inline_keyboard_markup=create_keyboard_from_list(markup_group_reopen_list)
                )

            # Удаляем сообщение с кнопками управления у оператора
            await bot.delete_messages(chat_id=operator_user_id, message_ids=[event.callback_query.message.message_id])
            await bot.answer_callback_query(query_id=event.callback_query.query_id, text="Вы отказались от заявки.")
            last_press.pop(operator_user_id, None) # Сброс состояния, если было

        elif action == 'Закрыть заявку':
            ReqNumb = int(callback_parts[1])
            message_in_group_id = callback_parts[2]

            req = db_roBOD['requests'].find_one_and_update(
                {'ReqNumb': ReqNumb, 'ReqOperatorUserId': operator_user_id},
                {"$set": {'ReqExecEndDt': datetime.now(), 'Status': 'Выполнено'}},
                return_document=ReturnDocument.AFTER # Получаем обновленный документ
            )

            if not req:
                logger.warning(f"Попытка закрыть заявку {ReqNumb}, которая не принадлежит оператору {operator_user_id} или не найдена.")
                await bot.answer_callback_query(query_id=event.callback_query.query_id, text="Не удалось закрыть заявку.", show_alert=True)
                return

            logger.info(f"Заявка {ReqNumb} закрыта оператором {operator_user_id}")
            Answer = req.get('Answer', "Ответ не предоставлен.")

            # Обновляем сообщение у оператора (убираем кнопки)
            msgtxt_operator_closed, _ = get_ReqText_from_db(ReqNumb)
            opertxt_closed = '\n\n<b>Выполнено</b>'
            await bot.edit_text(
                chat_id=operator_user_id,
                message_id=event.callback_query.message.message_id,
                text=msgtxt_operator_closed + opertxt_closed,
                parse_mode=ParseMode.HTML,
                inline_keyboard_markup= InlineKeyboardMarkup(inline_keyboard=[]) # Пустая клавиатура
            )

            # Обновляем сообщение в группе
            msgtxt_group_closed, _ = get_ReqText_from_db(ReqNumb) # Получаем текст без статуса
            text_for_group_closed = (f"{msgtxt_group_closed}\n\n"
                                     f"Исполнитель: @[{operator_user_id}].\n"
                                     f"Ответ:\n{Answer}\n\n"
                                     f"<b>Выполнено</b>")
            await bot.edit_text(
                chat_id=CHATID,
                message_id=message_in_group_id,
                text=text_for_group_closed,
                parse_mode=ParseMode.HTML,
                inline_keyboard_markup= InlineKeyboardMarkup(inline_keyboard=[]) # Пустая клавиатура
            )

            # Сообщаем пользователю-создателю, что заявка выполнена
            creator_user_id = req['CreatorUserId']
            if creator_user_id:
                msgtxt_to_creator_final = (f"Ответ по вашей заявке №<b>{ReqNumb}</b>:\n"
                                           f"<b>{Answer}</b>\n\n"
                                           f"Заявка исполнена.\n\n"
                                           f"Если возникли вопросы, свяжитесь с исполнителем @[{operator_user_id}].\nСпасибо.")
                await bot.send_text(chat_id=creator_user_id, text=msgtxt_to_creator_final, parse_mode=ParseMode.HTML)
            await bot.answer_callback_query(query_id=event.callback_query.query_id, text="Заявка закрыта.")
            last_press.pop(operator_user_id, None)

        elif action == 'Написать/Редактировать ответ':
            ReqNumb = int(callback_parts[1])
            message_in_group_id = callback_parts[2]
            last_press[operator_user_id] = [action, ReqNumb, message_in_group_id] # Сохраняем состояние
            await bot.send_text(
                chat_id=operator_user_id,
                text='Напишите, пожалуйста, ответ в одном сообщении. Далее необходимо нажать "Закрыть заявку", чтобы сообщение отправилось пользователю.',
                parse_mode=ParseMode.HTML
            )
            await bot.answer_callback_query(query_id=event.callback_query.query_id) # Просто подтверждаем получение колбэка

    except ValueError as ve:
        logger.error(f"Ошибка преобразования данных в callback: {callback_data_str} - {ve}")
        await bot.answer_callback_query(query_id=event.callback_query.query_id, text="Ошибка обработки данных.", show_alert=True)
    except Exception as e:
        logger.exception(f"Непредвиденная ошибка при обработке кнопки '{callback_data_str}': {e}")
        try:
            await bot.answer_callback_query(query_id=event.callback_query.query_id, text="Произошла ошибка.", show_alert=True)
        except Exception as e_ans:
            logger.error(f"Не удалось даже отправить answer_callback_query: {e_ans}")


async def main():
    # Регистрация обработчиков
    dispatcher.add_handler(CommandHandler(command="start", callback=start_command_cb))
    dispatcher.add_handler(CommandHandler(command="report", callback=report_command_cb))
    dispatcher.add_handler(MessageHandler(callback=message_cb)) # Для обычных текстовых сообщений
    dispatcher.add_handler(BotButtonCommandHandler(callback=buttons_answer_cb)) # Для нажатий на кнопки

    logger.info("Бот запускается...")
    try:
        await app.start_polling(dispatcher=dispatcher) # Явно передаем dispatcher
    except Exception as e:
        logger.critical(f"Критическая ошибка при запуске поллинга: {e}", exc_info=True)
    finally:
        logger.info("Бот останавливается...")
        client.close()
        logger.info("Соединение с MongoDB закрыто.")

if __name__ == "__main__":
    asyncio.run(main())