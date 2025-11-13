#GIT_VERSION=1.0
import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from bot.bot import Bot
from BotBase import BotBase
from DBManager import MongoDBManager
from MsgServices import MsgServices
from OthServices import OthServices
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
import requests

class EmergAlertBot (BotBase):
    def __init__(self):
        self.BotName = 'EmergAlertBot'
        # self.TOKEN = os.environ["TOKEN"]
        # self.TOKEN = os.environ["TOKEN_HELP"]
        self.TOKEN = os.environ["TOKEN_CHS"]
        self.bot = Bot(token=self.TOKEN, api_url_base=os.environ["API_URL_BASE"])
        self.db_manager = MongoDBManager()
        self.setup_handlers()
        self.MsgServ = MsgServices(self.BotName, self.TOKEN)
        self.OthServ = OthServices()
        #self.path = 'Risks.json'
        # self.path = 'data4.json'
        self.db_manager.alert_add_role('moseevay@veb.ru', 'Администратор', '')
        # self.db_manager.alert_add_role("sarumyanea@veb.ru", 'Администратор', '')
    # обработка start
    def start_cb(self, bot, event):
        print(self.bot.uin, event)
        userId = event.data['from']['userId'] 
        self.db_manager.put_status(self.BotName, event.data['from']['userId'], 'last_press', 'Начать')
        self.MsgServ.del_old_msgId(event)
        self.MsgServ.get_start(event)
        txt, markup = self.start_data(event)
        self.MsgServ.send_msg2(userId, txt, markup, False)
    # данные, если нажато вначало
    def start_data(self, event):
        userId = event.data['from']['userId'] 
        roles = self.db_manager.alert_get_roles(userId)
        fullname = f"{event.data['from']['lastName']} {event.data['from']['firstName']}"
        # получаем активный алерт, если он существует
        alert = self.db_manager.alert_get_active_alert()
        # пользователь подключился к боту
        txt1 = ''
        if alert:
            self.db_manager.alert_connected(alert['_id'], userId)
            # если есть ответ, то сообщаем об ответе + нужно добавить возможность изменить ответ
            Response = self.db_manager.alert_get_alert_response(userId, alert['_id'])
            if Response != 'Нет ответа' or len(alert['responses']) == 0:
                txt = '<b>\"' + alert['name'] + '\"\n\n </b>' + 'Ваш ответ \"' + Response + '\"'
                markup = [[{"text": 'Изменить ответ', "callbackData": 'Изменить ответ', "style": 'primary'}]]
                if 'Администратор' in roles:
                    txt1 += 'У Вас есть роль <b>Администратор</b>, Вы можете управлять оповещениями, а также получать отчеты по всем пользователям.\n\n'
                if 'КоординаторБлока' in roles:
                    role_data = self.db_manager.alert_get_role_data('КоординаторБлока', userId)
                    txt1 += 'У Вас есть роль <b>Координатор Блока \"' + role_data['block_name'] + '\"</b>, Вы отвечаете за предоставление информации в Блоке.\n\n'
                if 'Координатор' in roles:
                    role_data = self.db_manager.alert_get_role_data('Координатор', userId)
                    txt1 += 'У Вас есть роль <b>Координатор Группы \"' + role_data['group_name']  + '\"</b>, Вы отвечаете за предоставление информации в Группе\n\n'
                if txt1 != '':
                    txt1 += '\n'
                    # txt1 += 'Нажмите \"Опции\"'
                    # self.MsgServ.send_msg2(userId, txt1, '', False)
            else:
                txt, markup = self.alert_template(alert)
            if len(roles) > 1 and self.db_manager.alert_get_alert_response(userId, alert['_id']) != 'Нет ответа':
                markup.append([{"text": 'Опции', "callbackData": 'Опции', "style": 'attention'}])
            # self.MsgServ.send_msg2(userId, txt, '', False)
        elif 'Администратор' in roles:
            markup = [[{"text": 'Создать оповещение', "callbackData": 'Создать оповещение', "style": 'primary'}]]
            markup.append([{"text": 'Выгрузить пользователей в xlsx', "callbackData": 'Выгрузить пользователей в xlsx', "style": 'primary'}])
            markup.append([{"text": 'Загрузить/удалить пользователей', "callbackData": 'Загрузить/удалить пользователей', "style": 'primary'}])
            txt = 'Вы являетесь <b>Администратором</b> данного бота и можете создавать/удалять оповещения.'
        else:
            markup = ''
            txt = f'{fullname}, вы зарегистрированы в боте и будете получать оповещения!' # + '\nА также можете сообщить о ЧС, ответственным сотрудникам:'
            # markup = [[{"text": 'Сообщить о ЧС', "callbackData": 'Сообщить о ЧС', "style": 'primary'}]]
        return txt1 + txt, markup
    # Обработка кнопок (не url!)
    def buttons_answer_cb(self, bot, event):
        print(self.bot.uin, event)
        if event.data['callbackData'] != 'не реагировать':
            callbackData = re.split(';', event.data['callbackData'])
            self.db_manager.put_status(self.BotName, event.data['from']['userId'], 'last_press', event.data['callbackData'])
            # self.MsgServ.del_old_msgId(event) 
            userId = event.data['from']['userId'] 
            # role = self.db_manager.alert_get_role(userId)
            roles = self.db_manager.alert_get_roles(userId)
            alert = self.db_manager.alert_get_active_alert()
        # Начать эквивалентна /start
        if event.data['callbackData'] == 'Начать':
            self.start_cb(bot, event)
        # Действия для того, кто создает оповещение
        elif event.data['callbackData'] == 'Создать оповещение' and 'Администратор' in roles:
            # print(self.users)
            # self.MsgServ.get_start(event)
            # Возможно надо создать параметры по умолчанию и кнопки для ввода параметров оповещения
            txt = 'Введите название оповещения, используя буквы, цифры и знаки препинания:'
            #self.MsgServ.send_msg(event, txt, '', False)
            self.MsgServ.edit_msg(event, txt, '')
        # Загрузить/удалить оповещения
        elif event.data['callbackData'] == 'Загрузить/удалить пользователей' and 'Администратор' in roles:
            # print(self.users)
            self.MsgServ.get_start(event)
            # Возможно надо создать параметры по умолчанию и кнопки для ввода параметров оповещения
            txt = 'Загрузите xlsx файл с пользователями, по которым необходимо внести изменения:'
            self.MsgServ.send_msg(event, txt, '', False)
        # Добавить ответ
        elif event.data['callbackData'] == 'Добавить ответ' and 'Администратор' in roles:
            self.MsgServ.get_start(event)
            # Возможно надо создать параметры по умолчанию и кнопки для ввода параметров оповещения
            txt = 'Введите ответ для оповещения, используя буквы, цифры и знаки препинания.\nЧтобы ответ полностью уместился на кнопке, длина ответа должна быть не более 20 символов.'
            self.MsgServ.send_msg(event, txt, '', False)
        # Cообщить о ЧС
        elif event.data['callbackData'] == 'Сообщить о ЧС' and 'Администратор' in roles:
            self.MsgServ.get_start(event)
            # Возможно надо создать параметры по умолчанию и кнопки для ввода параметров оповещения
            txt = 'Введите сообщение о ЧС или о подозрении на ЧС. Ответственные оперативно изучат полученную информацию и примут решение об оповещении сотрудников.'
            self.MsgServ.send_msg(event, txt, '', False)
        # Выгрузка пользователей
        elif event.data['callbackData'] == 'Выгрузить пользователей в xlsx' and 'Администратор' in roles:
            self.MsgServ.get_start(event)
            # Список пользователей 
            file_stream = self.create_xlsx_users_report()
            bot.send_file(chat_id=userId, file=file_stream, caption="Список пользователей")
        # Запускаются оповещения
        elif event.data['callbackData'] == 'Запуск оповещения' and 'Администратор' in roles:
            self.db_manager.alert_add_alert(userId)
            alert = self.db_manager.alert_get_active_alert()
            # Список пользователей для запуска оповещения
            self.db_manager.alert_upload_users(alert['_id'], userId)
            # непосредственно отправка оповещения
            self.send_alert_to_users(alert)
        # Нажата кнопка опции
        elif event.data['callbackData'] == 'Опции' and len(roles) > 1:
            # self.MsgServ.get_start(event)
            markup = []
            if 'Администратор' in roles:
                markup.append([{"text": 'Завершить оповещение', "callbackData": 'Удалить оповещение', "style": 'primary'}])
                markup.append([{"text": 'Быстрый отчет', "callbackData": 'Быстрый отчет', "style": 'primary'}])
                markup.append([{"text": 'Полный отчет xlsx', "callbackData": 'Полный отчет xlsx', "style": 'primary'}])
                #markup.append([{"text": 'Назначить роль', "callbackData": 'Назначить роль', "style": 'primary'}])
            if 'КоординаторБлока' in roles:
                markup.append([{"text": 'Быстрый отчет по Блоку', "callbackData": 'Быстрый отчет по Блоку', "style": 'primary'}])
                markup.append([{"text": 'Отчет по Блоку xlsx', "callbackData": 'Отчет по Блоку xlsx', "style": 'primary'}])
                markup.append([{"text": 'Стать КоордГруппы', "callbackData": 'Назначить роль Координатора', "style": 'primary'}])
            if 'Координатор' in roles:
                markup.append([{"text": 'Быстрый отчет по Группе', "callbackData": 'Быстрый отчет по Группе', "style": 'primary'}])
                markup.append([{"text": 'Отчет по Группе xlsx', "callbackData": 'Отчет по Группе xlsx', "style": 'primary'}])
                markup.append([{"text": 'Заполнить за коллег', "callbackData": 'Заполнить данные за коллег', "style": 'primary'}])
            markup.append([{"text": 'Назад', "callbackData": 'Назад', "style": 'attention'}])
            self.MsgServ.edit_msg(event, 'Опции для оповещения:\n <b>\"' + alert['name'] + '\"</b>', markup)
            #self.MsgServ.send_msg(event, 'Опции для оповещения:\n <b>\"' + alert['name'] + '\"</b>', markup, False)
        # Опция удалить оповещение
        elif event.data['callbackData'] == 'Удалить оповещение' and 'Администратор' in roles:
            # self.MsgServ.get_start(event)
            alert = self.db_manager.alert_get_active_alert()
            self.send_alert_cancel_to_users(alert)
            self.db_manager.alert_deactivate(alert)
            self.MsgServ.send_msg(event, 'Оповещение удалено!', '', False)
            # список пользователей и ролей
            users_for_send = []
            order ={'Администратор': 1, 'КоординаторБлока': 2, 'Координатор': 3}
            for user in self.db_manager.alert_get_admins_and_coords():
                users_for_send.append([order[user['role']], user['role'], user['userId']])
            users_for_send.sort()
            # отправляем один отчет, если несколько ролей
            sended_users = set()
            for user in users_for_send:
                if user[2] not in sended_users:
                    file_stream = self.create_xlsx_report(alert['_id'], user[1], user[2])
                    bot.send_file(chat_id=user[2], file=file_stream, caption="Отчет по оповещению")
                    sended_users.add(user[2])
        # Быстрый отчет
        elif event.data['callbackData'] == 'Быстрый отчет' and 'Администратор' in roles:
            # self.MsgServ.get_start(event)
            txt, markup = self.get_fast_report(alert['_id'], 'Администратор', userId)
            markup = [[{"text": 'Назад', "callbackData": 'Опции', "style": 'attention'}]]
            # self.MsgServ.edit_msg(event, txt, markup)
            self.MsgServ.edit_msg(event, txt, markup)
        # Быстрый отчет по Блоку
        elif event.data['callbackData'] == 'Быстрый отчет по Блоку' and 'КоординаторБлока' in roles:
            #self.MsgServ.get_start(event)
            txt, markup = self.get_fast_report(alert['_id'], 'КоординаторБлока', userId)
            markup = [[{"text": 'Назад', "callbackData": 'Опции', "style": 'attention'}]]
            #self.MsgServ.edit_msg(event, txt, markup)
            self.MsgServ.edit_msg(event, txt, markup)
        # Быстрый отчет по группе
        elif event.data['callbackData'] == 'Быстрый отчет по Группе' and 'Координатор' in roles:
            #self.MsgServ.get_start(event)
            txt, markup = self.get_fast_report(alert['_id'], 'Координатор', userId)
            markup = [[{"text": 'Назад', "callbackData": 'Опции', "style": 'attention'}]]
            #self.MsgServ.edit_msg(event, txt, markup)
            self.MsgServ.edit_msg(event, txt, markup)
        # Отчет xlsx полный
        elif event.data['callbackData'] == 'Полный отчет xlsx' and 'Администратор' in roles:
            # self.MsgServ.get_start(event)
            file_stream = self.create_xlsx_report(alert['_id'], 'Администратор', userId)
            bot.send_file(chat_id=userId, file=file_stream, caption="Отчет по оповещению")
            #self.MsgServ.send_msg2(userId, file_stream, '', False)
        # Отчет по Блоку xlsx
        elif event.data['callbackData'] == 'Отчет по Блоку xlsx' and 'КоординаторБлока' in roles:
            # self.MsgServ.get_start(event)
            file_stream = self.create_xlsx_report(alert['_id'], 'КоординаторБлока', userId)
            bot.send_file(chat_id=userId, file=file_stream, caption="Отчет по оповещению")
        # Отчет по Группе xlsx 
        elif event.data['callbackData'] == 'Отчет по Группе xlsx' and 'Координатор' in roles:
            # self.MsgServ.get_start(event)
            file_stream = self.create_xlsx_report(alert['_id'], 'Координатор', userId)
            bot.send_file(chat_id=userId, file=file_stream, caption="Отчет по оповещению")
        # Назначить роль
        elif event.data['callbackData'] == 'Назначить роль' and 'Администратор' in roles:
            self.MsgServ.get_start(event)
            #self.get_non_respondents(userId, alert['_id'])
        # Изменить ответ пользователя
        elif event.data['callbackData'] == 'Заполнить данные за коллег' and 'Координатор' in roles:
            # self.MsgServ.get_start(event)
            # self.get_non_respondents(userId, alert['_id'])
            txt, markup = self.get_group_users(userId, alert['_id'])
            markup.append([{"text": 'Назад', "callbackData": 'Опции', "style": 'attention'}])
            self.MsgServ.edit_msg(event, txt, markup)
        # Нажат ответ на оповещение
        elif alert and event.data['callbackData'] in alert['responses']:
            # Добавляем ответ в БД
            self.db_manager.alert_add_response(alert['_id'], userId, event.data['callbackData'], userId)
            txt, markup = self.start_data(event)
            self.MsgServ.edit_msg(event, txt, markup)
        # Назад в главное меню
        elif alert and event.data['callbackData'] == 'Назад':
            txt, markup = self.start_data(event)
            self.MsgServ.edit_msg(event, txt, markup)
        # Изменить ответ
        elif event.data['callbackData'] == 'Изменить ответ':
            # Добавляем ответ в БД
            # self.MsgServ.get_start(event)
            txt, markup = self.alert_template(alert)
            self.MsgServ.edit_msg(event, txt, markup)
            # self.MsgServ.send_msg2(userId, txt, markup, False)
        # Выбран ответ за другого пользователя
        elif alert and callbackData[0] in alert['responses']:
            # Добавляем ответ в БД
            self.db_manager.alert_add_response(alert['_id'], callbackData[1], callbackData[0], userId)
            # Возврат в меню выбора ответов за пользователей
            # self.MsgServ.get_start(event)
            #self.get_non_respondents(userId, alert['_id'])
            txt, markup = self.get_group_users(userId, alert['_id'])
            markup.append([{"text": 'Назад', "callbackData": 'Опции', "style": 'attention'}])
            self.MsgServ.edit_msg(event, txt, markup)
            self.MsgServ.send_msg(event, 'Для @[' + callbackData[1] + '] Вами был выбран ответ <b>\"' + callbackData[0] + '\"</b>.', '', False)
            # Пользователю, за которого был проставлен ответ отправлено сообщение
            txt = 'За Вас был выбран ответ \"' +  callbackData[0] + '\" координатором @[' + userId + '].\nЕсли не согласны, то измените ответ.'
            try:
                self.MsgServ.send_msg2(callbackData[1], txt, '', False)
            except Exception as inst:
                print('Пользователь не подключен')
        # Назначить роль Координатора Группы
        elif 'КоординаторБлока' in roles and (event.data['callbackData'] in ('Назначить роль Координатора') 
                                              or callbackData[0] in ('Отключить роль КоордГруппы')
                                              or callbackData[0] in self.get_group_names(alert['_id'], self.db_manager.alert_get_role_data('КоординаторБлока', userId)['block_name'])) :
            if event.data['callbackData'] in ('Назначить роль Координатора'):
                txt = "Выберите группу, в которой у Вас будет роль Координатора:"
            elif callbackData[0] in ('Отключить роль КоордГруппы'):
                txt = "Выберите группу, в которой у Вас будет роль Координатора:"
                self.db_manager.alert_del_role(userId, 'Координатор')
            else:
                txt = 'Вы являетесь Координатором Группы <b>\"' + callbackData[0] + '\"</b>.\nВы можете стать Координатором другой группы или отказаться от данной роли:'
                self.db_manager.alert_add_role(userId, 'Координатор', callbackData[0])
            # self.MsgServ.get_start(event)
            role_data = self.db_manager.alert_get_role_data('КоординаторБлока', userId)
            role_data2 = self.db_manager.alert_get_role_data('Координатор', userId)
            curr_group = role_data2['group_name'] if role_data2 else ''
            # Выводим список групп
            groups = self.get_group_names(alert['_id'], role_data['block_name'])
            markup = []
            for group in groups:
                markup.append([{"text": group, "callbackData": group + ';' + userId if group != curr_group else 'не реагировать', "style": 'base' if group == curr_group else 'primary'}])
            markup.append([{"text": 'Откл. роль КоордГр', "callbackData": 'Отключить роль КоордГруппы;' + userId if curr_group != '' else 'не реагировать', "style": 'base' if curr_group == '' else 'primary'}])
            markup.append([{"text": 'Назад', "callbackData": 'Опции', "style": 'attention'}])
            self.MsgServ.edit_msg(event, txt, markup)
            # self.MsgServ.send_msg2(userId, txt, markup, False)
            #self.get_non_respondents(userId, alert['_id'])
    # Обработка сообщений
    def message_cb(self, bot, event):
        if not event.text.startswith('/') and event.data['chat']['type'] == "private":
            print(self.bot.uin, event)
            # self.MsgServ.del_old_msgId(event)
            userId = event.data['from']['userId'] 
            #role = self.db_manager.alert_get_role(userId)
            roles = self.db_manager.alert_get_roles(userId)
            last_press = re.split(';', self.db_manager.get_status(self.BotName, event.data['from']['userId'], 'last_press'))
            if last_press[0] == 'Создать оповещение' and 'Администратор' in roles:
                self.MsgServ.del_old_msgId(event)
                self.MsgServ.get_start(event)
                print(last_press[0])
                # Проверка введенного текста
                if not self.OthServ.is_text_valid(event.text):
                    txt = 'Текст не соответствует описанным критериям.\nНачните сначала создание оповещения.'
                    markup = ''
                else:
                    self.db_manager.alert_add_alert_tmp(userId, event.text, [])
                    txt = 'Предварительный просмотр оповещения:\n\n <b>\"'+ event.text + '\"</b>.\n\nДобавьте ответ или запустите оповещение:'
                    markup =[[{"text": 'Добавить ответ', "callbackData": 'Добавить ответ', "style": 'primary'}]]
                    markup.append([{"text": 'Запустить оповещение', "callbackData": 'Запуск оповещения', "style": 'attention'}])
                self.MsgServ.send_msg(event, txt, markup, False)
            elif last_press[0] == 'Добавить ответ' and 'Администратор' in roles:
                self.MsgServ.del_old_msgId(event)
                self.MsgServ.get_start(event)
                # Проверка введенного текста
                if not self.OthServ.is_text_valid(event.text):
                    txt = 'Текст не соответствует описанным критериям.\nНачните сначала создание оповещения.'
                    markup = ''
                else:
                    alert_tmp = self.db_manager.alert_add_response_in_alert_tmp(userId, event.text)
                    responses = alert_tmp['responses']
                    txt = 'Предварительный просмотр оповещения: <b>\n\n\"'+ alert_tmp['name'] + '\"</b>.'
                    markup = []
                    for i in range(len(responses)):
                        markup.append([{"text": responses[i], "callbackData": 'не реагировать', "style": 'base' }])
                    self.MsgServ.send_msg2(userId, txt, markup, False)
                    # Можно немного оптимизировать с предудыщим ответом
                    txt ='Добавьте ответ или запустите оповещение:'
                    markup =[[{"text": 'Добавить ответ', "callbackData": 'Добавить ответ', "style": 'primary'}]]
                    markup.append([{"text": 'Запустить оповещение', "callbackData": 'Запуск оповещения', "style": 'attention'}])
                self.MsgServ.send_msg(event, txt, markup, False)
            elif 'Администратор' in roles and last_press[0] == 'Загрузить/удалить пользователей':
                # загружаем файл 
                nm = event.data['from']['firstName'] + " " + event.data['from']['lastName']
                get_id = event.data['parts'][0]['payload']['fileId']  # получаем fileId
                get_file = bot.get_file_info(get_id).json()  # получаем json по файлу
                file_url = get_file['url']  # забираем значение url
                filename = get_file['filename']  # забираем значение filename
                response = requests.get(file_url)
                with open(filename, 'wb') as file:
                    file.write(response.content)  # сохраняем
                print(f"Приняли {filename} от {nm} через {file_url}")
                # создаем json
                self.xlsx_to_json(filename, 'data.json')
                # преобразование данных в словарь
                with open('data.json', 'r', encoding='utf-8') as file:
                    data = json.load(file)
                # загружаем в БД Монго
                self.db_manager.alert_upload_curr_users(userId, data)
                self.MsgServ.send_msg(event, 'Данные о пользователях обновлены.', '', False)
            elif last_press[0] in ('Начать', 'Сообщить о ЧС') and 'Администратор' not in roles:
                self.MsgServ.del_old_msgId(event)
                txt = 'Сообщение от ' + '@[' + userId + ']:\n' + event.text
                for admin in self.db_manager.alert_get_users_role('Администратор', ''):
                    self.bot.send_text(chat_id=admin, text=txt, parse_mode='HTML')
                    # self.MsgServ.send_msg2(admin, txt, '', False)
                self.MsgServ.get_start(event)
                txt = 'Спасибо! Ваше сообщение направлено Администраторам.'
                self.bot.send_text(chat_id=userId, text=txt, parse_mode='HTML')
                # self.MsgServ.send_msg2(userId, txt, '', False)
            elif 'Администратор' not in roles:
                txt = 'Если Вы хотите направить сообщение Администраторам бота, то нажмите кнопку \"В начало\", а потом пишите сообщение.'
                self.MsgServ.send_msg2(userId, txt, '', False)
    # Шаблон оповещений
    def alert_template(self, alert):
        responses = alert['responses']
        # txt, markup = self.alert_template(userId, alert['name'], alert['responses'], resp)
        txt = '<b>' + alert['name'] + '</b>'
        markup = []
        for i in range(len(responses)):
            markup.append([{"text": responses[i], "callbackData": responses[i], "style": 'primary' }])
        markup = '' if len(markup) == 0 else markup
        return txt, markup
    # Преобразование данных в словарь
    def upload_json(self, alertId, userId, path):
        # Чтение данных из JSON файла
        with open(path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        # Загрузка данных в БД
        print('Загрузка пользователей в БД')
    # Запуск оповещений пользователям
    def send_alert_to_users(self, alert):
        print('Запустили оповещения!')
        users = self.db_manager.alert_get_users(alert['_id'])
        for user in users:
            userId = user['user_email']
            txt, markup = self.alert_template(alert)
            print(userId)
            try:
                self.MsgServ.del_old_msgId2(userId)
                #if len(alert['responses']) == 0:
                self.MsgServ.get_start2(userId)
                self.MsgServ.send_msg2(userId, txt, markup, False)
                self.db_manager.alert_connected(alert['_id'], userId)
            except Exception as inst:
                print('Пользователь не подключен')
    # Отправка об отмене оповещения пользователям
    def send_alert_cancel_to_users(self, alert):
        users = self.db_manager.alert_get_users(alert['_id'])
        for user in users:
            userId = user['user_email']
            txt = 'Оповещение \"' + alert['name'] + '\" завершилось!\nВаш ответ - \"' + user["response"] + '\".'
            print(userId)
            try:
                self.MsgServ.del_old_msgId2(userId)
                self.MsgServ.get_start2(userId)
                # self.MsgServ.send_msg2(userId, txt, '', False)
                self.bot.send_text(chat_id=userId, text=txt, parse_mode='HTML')
            except Exception as inst:
                print('Пользователь не подключен')
    # Выгрузить отчет по пользователям в бот (вначале, кто не ответил)
    def get_fast_report(self, alertId, role, userId):
        # Данные роли
        role_data = self.db_manager.alert_get_role_data(role, userId)
        # Все пользователи
        users = self.db_manager.alert_get_users(alertId)
        # Варианты ответов активного алерта
        responses = self.db_manager.alert_get_active_alert()['responses']
        responses.append('Нет ответа') 
        # Создаем счетчики ответов
        responses_agr = {}
        for resp in responses:
            responses_agr[resp] = 0
        # формируем список пользователей для отчета
        response_lst = []
        responses_dict = {}
        # Администратор имеет доступ к ответам всех пользователей
        for user in users:
            # print(user['user_email'])
            resp = user['response']
            if (role == 'Координатор' and user['department_name'] == role_data['group_name']):
                response_lst.append([user['user_email'], resp])
                # responses_dict[user['user_email']] = resp
                responses_agr[resp] = responses_agr[resp] + 1
            elif (role == 'КоординаторБлока' and user['block_name'] == role_data['block_name']):
                if user['department_name'] not in responses_dict:
                    resp0 = {}
                    for r in responses:
                        resp0[r] = 0
                    responses_dict[user['department_name']] = resp0
                responses_dict[user['department_name']][resp] = responses_dict[user['department_name']][resp] + 1
                responses_agr[resp] = responses_agr[resp] + 1
            elif (role == 'Администратор'):
                if user['block_name'] not in responses_dict:
                    resp0 = {}
                    for r in responses:
                        resp0[r] = 0
                    responses_dict[user['block_name']] = resp0
                responses_dict[user['block_name']][resp] = responses_dict[user['block_name']][resp] + 1
                responses_agr[resp] = responses_agr[resp] + 1
        response_lst.sort(key=lambda x: (x[1], x[0]), reverse=False)
        # формируем отчет
        txt = '<b>Отчет по группе:</b>\n'
        if role == 'Администратор':
            txt = '<b>Полный отчет:</b>\n'
        elif role == 'КоординаторБлока':
            txt = '<b>Отчет по Блоку:</b>\n'
        # агр.данные
        for resp in responses:
            txt += resp + ' - ' + str(responses_agr[resp]) + '\n'
        txt += '\n'
        # для Координатора выводим всех коллег из группы с ответом
        if role == 'Координатор':
            for user in response_lst:
                resp = 'нет ответа' if user[1] == '' else user[1] 
                txt += '@[' + user[0] + ']' + ' - ' + resp + '\n'
        # для Админа и КоординатораБлока агрегированная инфа
        else:
            role_ = 'КоординаторБлока' if role == 'Администратор' else 'Координатор'
            for dep in responses_dict:
                # Список пользователей, соответствующих роли
                txt_role = 'Коодинаторы:\n'
                for user_role in self.db_manager.alert_get_users_role(role_, dep):
                    txt_role += '@[' + user_role + ']\n'
                txt += '<b>' + dep + '</b>\n' + txt_role
                for resp in responses:
                    txt += resp + ' - ' + str(responses_dict[dep][resp]) + '\n'
        return txt, ''
        # self.MsgServ.send_msg2(userId, txt, '', False)
    # Не ответившие (вывод всех в Группе)
    def get_non_respondents(self, userId, alertId):
        role = 'Координатор'
        # Данные роли
        role_data = self.db_manager.alert_get_role_data(role, userId)
        # Все пользователи
        users = self.db_manager.alert_get_users(alertId)
        # формируем список пользователей для отчета
        response_lst = []
        # Администратор имеет доступ к ответам всех пользователей
        for user in users:
            if ((role == 'Администратор') or (role == 'Координатор' and user['department_name'] == role_data['group_name']) or (role == 'КоординаторБлока' and user['block_name'] == role_data['block_name'])):
                response_lst.append([user['user_fullname'], user['user_email'], user['response']])
        # Сортировка
        response_lst = sorted(response_lst, key=lambda x: (0 if x[2] == 'Нет ответа' else 1, x[0]))
        # Ответы
        responses = self.db_manager.alert_get_active_alert()['responses']
        # формируем отчет
        txt = 'Заполните информацию за коллег:\n'
        self.MsgServ.send_msg2(userId, txt, '', False)
        for user in response_lst:
            markup = []
            for i in range(len(responses)):
                markup.append([{"text": responses[i], "callbackData": responses[i] + ';' + user[1] if responses[i] != user[2]  else 'не реагировать', "style": 'primary' if responses[i] != user[2]  else 'base' }])
            self.MsgServ.send_msg2(userId, '@[' + user[1] + ']' , markup, False)
    # Люди в группе с ответами
    def get_group_users(self, userId, alertId):
        role = 'Координатор'
        # Данные роли
        role_data = self.db_manager.alert_get_role_data(role, userId)
        # Все пользователи
        users = self.db_manager.alert_get_users(alertId)
        # формируем список пользователей для отчета
        response_lst = []
        # Администратор имеет доступ к ответам всех пользователей
        for user in users:
            if ((role == 'Администратор') or (role == 'Координатор' and user['department_name'] == role_data['group_name']) or (role == 'КоординаторБлока' and user['block_name'] == role_data['block_name'])):
                response_lst.append([user['user_fullname'], user['user_email'], user['response']])
        # Сортировка
        response_lst = sorted(response_lst, key=lambda x: (0 if x[2] == 'Нет ответа' else 1, x[0]))
        # Ответы
        responses = self.db_manager.alert_get_active_alert()['responses']
        # формируем отчет
        txt = 'Заполните информацию за коллег:\n'
        # self.MsgServ.send_msg2(userId, txt, '', False)
        markup = []
        for user in response_lst:
            markup.append([{"text": user[0], "callbackData": 'не реагировать', "style": 'base' }])
            for i in range(len(responses)):
                markup.append([{"text": responses[i], "callbackData": responses[i] + ';' + user[1] if responses[i] != user[2]  else 'не реагировать', "style": 'primary' if responses[i] != user[2]  else 'base' }])
        return txt, markup 
    # Создание xls отчета
    def create_xlsx_report(self, alertId, role, userId):
        wb = Workbook()
        ws = wb.active
        ws.append(
            ['Имя', 'Подключен к боту', 'Ответ', 'Время ответа', 'Координатор группы Имя', 
             'Группа', 'Координатор блока Имя', 'Блок', 'Кто заполнил email', 'Комментарий'])  # Добавляем заголовок для времени ответа
        '''
        ws.append(
            ['Пользователь email', 'Имя', 'Подключен к боту', 'Ответ', 'Время ответа', 'Координатор группы email', 'Координатор группы Имя', 
             'Группа', 'Координатор блока email', 'Координатор блока Имя', 'Блок', 'Кто заполнил email', 'Комментарий'])  # Добавляем заголовок для времени ответа
        '''
        # Получение списка пользователей
        users = self.db_manager.alert_get_users(alertId)
        role_data = self.db_manager.alert_get_role_data(role, userId)
        data = []
        for user in users:
            if (role == 'Администратор') or (role == 'Координатор' and user['department_name'] == role_data['group_name']) or (role == 'КоординаторБлока' and user['block_name'] == role_data['block_name']):
                '''
                data.append([user['user_email'], user['user_fullname'], user['is_connected'], user['response'], user['responseDate'], user['supervisor_email'], user['supervisor_fullname'], 
                             user['department_name'],user['block_coordinator_email'], user['block_coordinator_fullname'], user['block_name'], user['userId_response'], user['comment']])
                '''
                data.append([user['user_fullname'], user['is_connected'], user['response'], user['responseDate'], user['supervisor_fullname'], 
                             user['department_name'], user['block_coordinator_fullname'], user['block_name'], user['userId_response'], user['comment']])
        sorted_data = sorted(data, key=lambda x: (x[7], x[5], x[0]))  # 9 - индекс столбца "Блок", 6 - индекс столбца "Группа"
        for row in sorted_data:
            ws.append(row)
        '''
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in sorted_data:
            ws.append(row)
            if row[2] == "No response":
                for cell in ws[ws.max_row]:
                    cell.fill = red_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        ws.auto_filter.ref = ws.dimensions
        '''
        if role == 'Администратор':
            report_name = 'полный отчет.xlsx'
        elif role == 'КоординаторБлока':
            report_name = 'отчет по Блоку.xlsx'
        elif role == 'Координатор':
            report_name = 'отчет по Группе.xlsx'

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)  # Возвращаем курсор в начало файла
        report_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        file_stream.name = f'{report_timestamp} ' + report_name
        return file_stream
    # Создание xls отчета
    def create_xlsx_users_report(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['user_email', 'user_fullname', 'supervisor_email', 'supervisor_fullname',
                   'block_coordinator_email', 'block_coordinator_fullname', 'department_name', 'block_name', 'admin_email'])  # Добавляем заголовок для времени ответа
        # Получение списка пользователей
        users = self.db_manager.alert_get_curr_users()
        data = []
        for user in users:
            data.append([user['user_email'], user['user_fullname'], user['supervisor_email'], user['supervisor_fullname'],
                         user['block_coordinator_email'], user['block_coordinator_fullname'], user['department_name'], user['block_name'], ''])
        # print(data)
        sorted_data = sorted(data, key=lambda x: (x[6], x[7], x[0]))  # 9 - индекс столбца "Блок", 6 - индекс столбца "Группа"
        for row in sorted_data:
            ws.append(row)

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)  # Возвращаем курсор в начало файла
        report_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        file_stream.name = f'{report_timestamp} ' + 'пользователи.xlsx'
        return file_stream
    # Блоки
    def get_block_names(self, alertId):
        users = self.db_manager.alert_get_users(alertId)
        blocks = set()
        for user in users:
            blocks.add(user['block_name'])
        blocks = list(blocks)
        return blocks.sort()
    # Группы
    def get_group_names(self, alertId, block_name): 
        users = self.db_manager.alert_get_users(alertId)
        groups = set()
        for user in users:
            if user['block_name'] == block_name:
                groups.add(user['department_name'])
        groups = list(groups)
        groups.sort()
        return groups
    # xls to json
    def xlsx_to_json(self, xlsx_file, json_file):
        # Читаем файл Excel
        df = pd.read_excel(xlsx_file)
        # Заменяем значения NaN на пустые строки
        df = df.fillna('')
        # Преобразуем DataFrame в JSON
        json_data = df.to_json(orient='records', force_ascii=False)
        # Сохраняем JSON в файл
        with open(json_file, 'w', encoding='utf-8') as f:
            f.write(json_data)
if __name__ == "__main__":
    bot_app = EmergAlertBot()
    bot_app.run()
