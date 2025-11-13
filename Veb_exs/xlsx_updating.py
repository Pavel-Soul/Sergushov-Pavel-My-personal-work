import openpyxl

def update_descriptions(source_file, source_sheet, target_file, target_sheet, save_as_new_file=True):
    print(f"Загрузка файлов '{source_file}' и '{target_file}'")
    try:
        source_workbook = openpyxl.load_workbook(source_file)
        target_workbook = openpyxl.load_workbook(target_file)
    except FileNotFoundError as e:
        print(f"Ошибка: {e}")
        return

    source_sheet = source_workbook[source_sheet]
    target_sheet = target_workbook[target_sheet]

    # Считываем данные из файла 
    descriptions = {}
    for row in source_sheet.iter_rows(min_row=3, max_row=source_sheet.max_row, min_col=1, max_col=3):
        name = row[0].value  # Название нейросети из столбца A
        description = row[2].value  # Описание из столбца C
        if name and description:
            descriptions[name.strip()] = description.strip()
    print(f"Найдено {len(descriptions)} описаний для обновления.")

    # Обновляем данные в целевом файле 
    updated_count = 0
    unupdated_count = 0
    for row in target_sheet.iter_rows(min_row=3, max_row=target_sheet.max_row, min_col=1, max_col=target_sheet.max_column):
        name = row[1].value  # Название нейросети из столбца B
        description_cell = row[7]  # Описание в столбце H

        if name and name.strip() in descriptions:
            description_cell.value = descriptions[name.strip()]
            updated_count += 1
            print(f"Обновлено описание для: {name.strip()}")
        else:
            unupdated_count += 1
            #print(f"Не найдено описание для: {name.strip() if name else 'пустое имя'}")
    print(f"Обновлено {updated_count} описаний.")

    # Сохранение результата
    if save_as_new_file:
        output_filename = target_file.replace(".xlsx", "_обновлено.xlsx")
        target_workbook.save(output_filename)
        print(f"Изменения сохранены в новом файле: {output_filename}")
    else:
        target_workbook.save(target_file)
        print(f"Изменения сохранены в исходном файле: {target_file}")

if __name__ == "__main__":
    source_file = "Нейросети_16_07_2025.xlsx"
    source_sheet = "Лист1"
    target_file = "Кнопки ServiceBot_llms.xlsx"
    target_sheet = "Лист1"

    # save_as_new_file=False, чтобы сохранить изменения в исходном файле
    update_descriptions(source_file, source_sheet, target_file, target_sheet, save_as_new_file=True)